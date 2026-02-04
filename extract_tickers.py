#!/usr/bin/env python3
"""
Bootstrap utility: reads ATLAS_data.xlsx and writes ticker universe + parameters
into config/atlas_config.yaml for use by bloomberg_loader.py.
"""

import argparse
import os
import sys

import openpyxl
import yaml


def extract_config(
    xlsx_path: str,
    output_yaml: str,
    output_xlsx: str | None = None,
) -> None:
    print(f"Opening {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- parameters sheet ---
    ws_params = wb["parameters"]
    params: dict = {}
    for row in ws_params.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
        key, val = row
        if key is None:
            continue
        key_lower = str(key).strip().lower().replace(" ", "_")
        if hasattr(val, "strftime"):
            val = val.strftime("%Y-%m-%d")
        params[key_lower] = val

    start_date = params.get("start_date", "2013-01-01")
    end_date = params.get("end_date", "2026-02-03")
    period = params.get("period", "D")
    currency = params.get("currency", "EUR")

    # --- tickers from price sheet header row ---
    ws_price = wb["price"]
    header_row = next(ws_price.iter_rows(min_row=1, max_row=1, values_only=True))
    tickers = [str(v).strip() for v in header_row if v is not None and str(v).strip() != "Ticker"]

    print(f"  Found {len(tickers)} tickers")
    print(f"  Date range: {start_date} -> {end_date}")
    print(f"  Period: {period}, Currency: {currency}")

    wb.close()

    # --- fields (matching the static sheet names in the xlsx) ---
    fields = {
        "price": "PX_LAST",
        "Pxtobook": "PX_TO_BOOK_RATIO",
        "EPS": "IS_EPS",
    }

    # --- assemble config ---
    if output_xlsx is None:
        base, ext = os.path.splitext(xlsx_path)
        output_xlsx = base + "_static" + ext

    config = {
        "parameters": {
            "start_date": start_date,
            "end_date": end_date,
            "period": period,
            "currency": currency,
        },
        "paths": {
            "source_xlsx": xlsx_path,
            "output_xlsx": output_xlsx,
        },
        "bloomberg": {
            "batch_size": 250,
            "ticker_suffix": " Equity",
        },
        "fields": fields,
        "tickers": tickers,
    }

    os.makedirs(os.path.dirname(output_yaml) or ".", exist_ok=True)
    with open(output_yaml, "w") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

    print(f"  Wrote config to {output_yaml}")


def main():
    parser = argparse.ArgumentParser(description="Extract tickers from ATLAS_data.xlsx into YAML config")
    parser.add_argument(
        "--xlsx",
        default=r"X:\Quant\Data\ATLAS_data.xlsx",
        help="Path to source ATLAS_data.xlsx",
    )
    parser.add_argument(
        "--output-yaml",
        default=os.path.join(os.path.dirname(__file__), "config", "atlas_config.yaml"),
        help="Output YAML config path",
    )
    parser.add_argument(
        "--output-xlsx",
        default=None,
        help="Output xlsx path (default: <input>_static.xlsx)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"ERROR: source file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    extract_config(args.xlsx, args.output_yaml, args.output_xlsx)


if __name__ == "__main__":
    main()
